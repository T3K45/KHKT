import PySimpleGUI as sg
import pandas as pd
import yagmail
import os
import requests
import webbrowser
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
from bs4 import BeautifulSoup
import wmi
import sys

dong1 = os.path.dirname(os.path.abspath(__file__))
dong2 = "1.xlsx"
dong3 = "XepLoai.xlsx"


def get_serial_number():
    try:
        wmi_obj = wmi.WMI()
        system_serial = wmi_obj.Win32_BIOS()[0].SerialNumber
        return system_serial
    except Exception as e:
        sg.popup(f"Không thể lấy serial number: {str(e)}")
my_seri = "Seri Number : " + get_serial_number()
url = 'https://tonghopvaquanlyloivipham.nguyendinhchien.io.vn/'
response = requests.get(url)
seri = []
new_public = []
if response.status_code == 200:
    html_content = response.text
    soup = BeautifulSoup(html_content, 'html.parser')
    elements_with_id = soup.find_all(id=["seri","trangthai"])
    for element in elements_with_id:
        if element.get('id').strip():
            text_content = element.get_text()
            if 'Seri Number' in text_content and text_content != 'Seri Number : ':
                seri.append(text_content)
            if element.get('id') == 'trangthai':
                public = text_content.split(':')[1].strip()
                new_public.append(public)

def start():
    def sendmail(tieude,noidung,file):
        fin = open("mailltruongthptdl3.txt","r",encoding="utf8")
        mailsent = fin.readline().strip("\n")
        matkhau = fin.readline()
        yag = yagmail.SMTP(f"{mailsent}",f"{matkhau}")
        fout = open("danhsachmailgvcn.txt","r",encoding="utf8")
        email_sent = []
        mail = fout.read().split()
        email_sent.append(mail)
        for to_email in email_sent:
            yag.send(to=to_email,subject=tieude,contents=noidung,attachments=file)

    def contact():
        link = "https://www.facebook.com/ngv.khanh.nlqda"
        link1 = "https://www.facebook.com/dinhchien85"
        webbrowser.open_new_tab(link)
        webbrowser.open_new_tab(link1)

    def xoatatca(file_path, sheet_name, start_row=2):
        try:
            workbook = load_workbook(filename=file_path)
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                if start_row >= 2:
                    for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                        for cell in row:
                            cell.value = None
                    workbook.save(file_path)
        except Exception as e:
            sg.popup_error(f"Lỗi xảy ra: {str(e)}",icon="logo.ico")

    def xoainput():
        for key in values:
            window[key](" ")
        return None

    def docfile():
        directory_path = dong1
        if os.path.exists(directory_path) and os.path.isdir(directory_path):
            file_list = os.listdir(directory_path)
            excel_files = [file for file in file_list if file.endswith('.xlsx')]
            return excel_files
        else:
            sg.popup("Đường dẫn không tồn tại hoặc không phải là thư mục.Xem Lại Phần setup.txt",icon="logo.ico")
            return None 
            
    def hienthi_dulieu(table_data_result, table_header_result):
        layout_result = [
            [sg.Table(values=table_data_result, headings=table_header_result, key="TableResult",
                    row_height=40, justification="center", expand_x=True, expand_y=True)],
            [sg.Button("Đóng", key="Dong", size=(15), button_color="red")]
        ]

        window_result = sg.Window("Danh Sách Học Sinh Vi Phạm Lỗi Tuần "+f"{tuan}", layout_result, icon="logo.ico")

        while True:
            event_result, values_result = window_result.read()

            if event_result in (sg.WINDOW_CLOSED, "Dong"):
                break

        window_result.close()
        
    def timkiem(name_file, name_sheet):
        while True:
            all_sheets = pd.ExcelFile(name_file).sheet_names
            if name_sheet not in all_sheets:
                sg.popup_error(f"{name_sheet }" + " không tồn tại trong cơ sở dữ liệu.", icon="logo.ico")
                if not name_sheet:
                    break
                sg.popup("Truy Cập Thất Bại", icon="logo.ico")
                break
            else:
                df = pd.read_excel(name_file, sheet_name=name_sheet)
                table_data_result = df.values.tolist()
                table_header_result = df.columns.values.tolist()
                hienthi_dulieu(table_data_result, table_header_result)
                sg.popup("Truy Cập Thành Công", icon="logo.ico")
                break
            
    def xeploai(input_file_path, output_file_path, sheet_name):
        try:
            data = pd.read_excel(input_file_path, sheet_name=sheet_name)
            df = pd.DataFrame(data)
            df['Tổng Điểm Trừ'] = df.groupby('Lớp')['Điểm Trừ'].transform('sum')
            result = df.drop_duplicates(subset='Lớp')[['Lớp', 'Tổng Điểm Trừ']].sort_values(by='Tổng Điểm Trừ', ascending=False)
            result['Tổng Điểm Trừ'] = result['Tổng Điểm Trừ'].fillna(0)
            result['Tổng Điểm Trừ'] = pd.to_numeric(result['Tổng Điểm Trừ'], errors='coerce')
            result['Tổng Điểm Trừ'] = result['Tổng Điểm Trừ'].fillna(0)
            result['Xếp Thứ'] = result['Tổng Điểm Trừ'].rank(method='dense', ascending=False).astype(int)
            output_path = Path(output_file_path)
            if not output_path.is_file():
                result.to_excel(output_file_path, index=False)
            book = load_workbook(output_file_path)
            if sheet_name in book.sheetnames:
                sheet = book[sheet_name]
                for r_idx, row in enumerate(dataframe_to_rows(result, index=False, header=True), sheet.min_row):
                    for c_idx, value in enumerate(row, sheet.min_column):
                        sheet.cell(row=r_idx, column=c_idx, value=value)
                book.save(output_file_path)
        except Exception as e:
            sg.popup_error(f"Lỗi xảy ra: {str(e)}",icon="logo.ico")

    def savefile(data,tuan):
        df = pd.DataFrame(data)
        sheet_name = 'Tuần '+f"{tuan}" 
        output_file = f"{dong2}"
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            sg.popup("Lưu File Thành Công",icon="logo.ico")
        except:
            sg.popup_error("Tên Tuần Trùng Với CSDL Đã Lưu Trước Đó",icon="logo.ico")

    def savefilexeploai(data,tuan):
        df = pd.DataFrame(data)
        sheet_name = 'Tuần '+f"{tuan}" 
        output_file = f"{dong3}"
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        except:
            return

    def them_du_lieu(values, dong2):
        try:
            book = load_workbook(dong2)
            sheet = book['Sheet1']
            data = sheet.values
            columns = next(data)
            df = pd.DataFrame(data, columns=columns)
            new_row = pd.DataFrame([values], columns=df.columns)
            df = pd.concat([df, new_row], ignore_index=True)
            df['STT'] = range(1, len(df) + 1)
            for row in sheet.iter_rows(min_row=2, min_col=1, max_row=sheet.max_row, max_col=sheet.max_column):
                for cell in row:
                    cell.value = None
            for r_idx, row in df.iterrows():
                for c_idx, value in enumerate(row):
                    sheet.cell(row=r_idx + 2, column=c_idx + 1, value=value)
            book.save(dong2)
            book.close()
            df = pd.read_excel(dong2, sheet_name='Sheet1')
            xoainput()
            new_values = df.values.tolist()
            window["Table"].update(values=new_values)
            sg.popup("Thêm Thành Công", icon="logo.ico")
        except Exception as e:
            sg.popup_error(f"Lỗi: {str(e)}",icon="logo.ico")

    def xoahang(dong2, selected_row):
        try:
            book = load_workbook(dong2)
            sheet = book['Sheet1']
            sheet.delete_rows(selected_row + 2)
            book.save(dong2)
            book.close()
        except Exception as e:
            sg.popup_error(f"Lỗi: {str(e)}",icon="logo.ico")

    def update(dong2, window):
        try:
            df = pd.read_excel(dong2, sheet_name='Sheet1')
            data = df.values.tolist()
            window['Table'].update(values=data)
        except Exception as e:
            sg.popup_error(f"Lỗi: {str(e)}",icon="logo.ico")

    def sapxeplop(file_path):
        try:
            book = load_workbook(file_path)
            if 'Sheet1' in book.sheetnames:
                sheet = book['Sheet1']
                data = pd.DataFrame(sheet.values)
                header = data.iloc[0]
                data = data[1:]
                data.columns = header
                data_sorted = data.sort_values(by='Lớp', ascending=True)
                data_sorted['STT'] = range(1, len(data_sorted) + 1)
                sheet.delete_rows(2, sheet.max_row)
                for row in data_sorted.values.tolist():
                    sheet.append(row)
                book.save(file_path)
                book.close()
            else:
                sg.popup("Không tìm thấy Sheet1 trong file Excel.",icon="logo.ico")
        except Exception as e:
            sg.popup_error(f"Lỗi: {str(e)}",icon="logo.ico")

    def khoangcach(file_path, increase_factor=3):
        workbook = load_workbook(filename=file_path)
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                ws.column_dimensions[column].width = max_length * increase_factor
        workbook.save(file_path)

    if os.path.isfile(dong2):
        df = pd.read_excel(dong2)
    else:
        data = {
            'STT': [],
            'Thời Gian': [],
            'Họ Và Tên': [],
            'Lớp': [],
            'Lỗi Vi Phạm': [],
            'Điểm Trừ': [],
        }
        df = pd.DataFrame(data)
        df.to_excel(dong2, index=False)

    if os.path.isfile(dong3):
        df = pd.read_excel(dong3)
    else:
        data = {
            'Lớp': [],
            'Tổng Điểm Trừ': [],
            'Xếp Thứ': []
        }
        df = pd.DataFrame(data)
        df.to_excel(dong3, index=False)
        
    khoangcach(dong2)
    khoangcach(dong3)
    df = pd.read_excel(rf"{dong2}")
    table_data = df.values.tolist()
    table_header = df.columns.values.tolist()
    sg.theme("BlueMono")
    layout1 = [
        [sg.Text("DANH SÁCH HỌC SINH VI PHẠM NỘI QUY",background_color="Green",text_color="Yellow",justification="center",size=(95))],
        [sg.Text("Ngày",size=(10,1)),
        sg.InputText(key='Thời Gian', size=(14,1)),sg.CalendarButton('', target='Thời Gian', key='-CALENDAR-', format='%d/%m/%Y',size=(2,1)),
        sg.Text("Lớp",size=(10,1)),sg.Combo(["10D1","10D2","10D3","10D4","10D5","10D6","10D7","10D8","10T1","10T2","10T3","10T4","10T5",
                                            "11A1","11D1","11D2","11D3","11D4","11D5","11D6","11D7","11T1","11T2","11T3","11T4","11T5",
                                            "12A1","12B","12D1","12D2","12D3","12D4","12D5","12D6","12D7","12T1","12T2","12T3","12T4"],key="Lớp",size=(13,7)),
        sg.Text("Điểm Trừ",size=(10,1)),sg.Input(key="Điểm Trừ",size=(15,1))],
        [sg.Text("Họ Và Tên",size=(10,1)),sg.Input(key="Họ Và Tên",size=(80,4))],
        [sg.Text("Lỗi Vi Phạm",size=(10,1)),sg.Multiline(key="Lỗi Vi Phạm",size=(80,3))]
                ]

    layout2 =[
        [sg.Button("Thêm Học Sinh Vi Phạm",button_color="Red",key="Thêm",size=(23,1)),
        sg.Button("Xếp Lại Lớp",key="xeplop",size=(23,1)),
        sg.Button("Xử Lý Thông Tin",key="xuly",size=(23,1)),
        sg.Button("Lưu File",key="Save",button_color="Red",size=(23,1))],
        
        [sg.Button("Xoá Hàng",key="delete",size=(23,1)),
        sg.Button("Xoá Tất Cả",key="reset",size=(23,1),button_color="orange"),
        sg.Button("Tìm Kiếm Danh Sách Lỗi Tuần",key="find",button_color="orange",size=(23,1)),
        sg.Button("Tìm Kiếm Xếp Loại Tuần",key="find_xeploai",size=(23,1))],
        
        [sg.Button("Contact",key="contact",button_color="red",size=(23,1)),
        sg.Button("Gửi GVCN",button_color="Green",key="sendmail",size=(23,1)),
        sg.Button("Help",button_color="Green",key="help",size=(23,1)),
        sg.Button("Thoát Phầm Mềm",key="Exit",button_color="red",size=(23,1))],          
                ]

    image = [
        [sg.Image("logo.png")]
                ]

    layout = [
        [sg.Frame("THPT ĐÔ LƯƠNG III",image,size=(135,155)),sg.Frame("Nhập Liệu",layout1,size=(650,155))],
        [sg.Frame("Công Cụ",layout2,size=(795,125))],
        [sg.Table(values= table_data,
                headings= table_header,
                key="Table",
                row_height = 35,
                justification="center",
                expand_x=True,
                expand_y=True,
                )]  
                ]

    window = sg.Window("PHẦN MỀM TỔNG HỢP VÀ QUẢN LÝ LỖI VI PHẠM - [ NHÓM TÁC GIẢ : NGUYỄN ĐÌNH CHIẾN - NGUYỄN VIẾT KHANH ]",layout,icon="logo.ico")
    while True:
        event,values = window.read()
        if event == sg.WIN_CLOSED or event == "Exit":
            break
        if event == "Thêm":
            them_du_lieu(values, dong2)

        if event == "delete":
            df = pd.read_excel(dong2, sheet_name=None)
            data = df['Sheet1'].values.tolist()
            selected_row = values['Table'][0] if values['Table'] else None
            if selected_row is not None:
                xoahang(dong2, selected_row)
                update(dong2, window)
                
        if event == "xuly":
            sheet_name = 'Sheet1'
            xeploai(dong2, dong3, sheet_name)
            sg.popup("Xử Lý Thành Công",icon="logo.ico")
            khoangcach(dong3)
            
        if event == "xeplop": 
            sapxeplop(dong2)
            sg.popup("Xếp Thành Công",icon="logo.ico")
            df = pd.read_excel(rf"{dong2}")
            df['STT'] = range(1, len(df) + 1)
            new_values = df.values.tolist()
            window["Table"].update(values=new_values)

        if event == "reset":
            sheet_can_xoa = 'Sheet1'
            hang_bat_dau_xoa = 2 
            xoatatca(dong2, sheet_can_xoa, hang_bat_dau_xoa)
            xoatatca(dong3, sheet_can_xoa, hang_bat_dau_xoa)
            sg.popup("Xoá Thành Công",icon="logo.ico")
            df = pd.read_excel(rf"{dong2}")
            new_values = df.values.tolist()
            window["Table"].update(values=new_values)

        if event == "contact":
            contact()
        if event == "help":
            layout_help = [
                [sg.Button("Chức Năng Các Nút Lệnh",size=(100,3),button_color="green")],
                [sg.Button("",size=(25,2),key="time_help"),
                sg.Button("Chọn Ngày Tháng Năm Theo Mẫu",size=(90,2))],
                [sg.Button("Thêm Học Sinh Vi Phạm",button_color="red",size=(25,2),key="add"),
                sg.Button("Thêm Học Sinh Bị Lỗi Vào Danh Sách",size=(90,2))],
                [sg.Button("Xếp Lại Lớp",size=(25,2),key="sap_xep"),
                sg.Button("Sắp Xếp Lại Các Lớp Theo Thứ Tự Từ Thấp Đến Cao",size=(90,2))],
                [sg.Button("Xử Lý Thông Tin",size=(25,2),key="xu_ly"),
                sg.Button("Xử Lý Các Thông Tin Nhập Từ Máy Tính Và Phân Loại Nó",size=(90,2))],
                [sg.Button("Lưu File",button_color="red",size=(25,2),key="luu_lai"),
                sg.Button("Lưu Lại Vào Cơ Sở Dữ Liệu Sau Khi Đã Xử Lý Hoàn Tất",size=(90,2))],
                [sg.Button("Xóa Hàng",size=(25,2),key="xoa_hang"),
                sg.Button("Xóa Vào Hàng Muốn Chọn ( Click Vào Hàng Muốn Xóa Trên Màn )",size=(90,2))],
                [sg.Button("Xóa Tất Cả",button_color="orange",size=(25,2),key="xoa_tatca"),
                sg.Button("Xóa Tất Cả Các Thông Tin Học Sinh/Lỗi Vi Phạm Hiển Thị Trên Màn Hình",size=(90,2))],
                [sg.Button("Tìm Kiếm Danh Sách Lỗi Tuần",button_color="orange",size=(25,2),key="ds_loivp"),
                sg.Button("Hiển Thị DS Lỗi Vi Phạm Đã Tổng Hợp Lại Về Một Tuần Nào Đó Được Lưu Lại Trong CSDL",size=(90,2))],
                [sg.Button("Tìm Kiếm Xếp Loại Tuần",size=(25,2),key="ds_xeploai"),
                sg.Button("Hiển Thị DS Xếp Loại Đã Tổng Hợp Lại Về Một Tuần Nào Đó Được Lưu Lại Trong CSDL",size=(90,2))],
                [sg.Button("Contact",size=(25,2),key="tacgia"),
                sg.Button("Tác Giả, Liên Hệ Xử Lý Khi Phầm Mềm Gặp Vẫn Đề",size=(90,2))],
                [sg.Button("Gửi GVCN",button_color="green",size=(25,2),key="gui_mail"),
                sg.Button("Gửi Về Mail của các Giáo Viên Trong Nhà Trường",size=(90,2))],
                [sg.Button("Thoát",button_color="red",size=(100,2),key="help_exit")]
            ]
            help_window = sg.Window("Chức Năng Các Nút Lệnh",layout_help,icon="logo.ico",size=(800,650))
            while True:
                help_event,help_value = help_window.read()
                if help_event in (sg.WINDOW_CLOSED, "help_exit"):
                    break
                if help_event == "time_help":
                    sg.popup("Chọn Ngày Tháng Năm Theo Mẫu")
                if help_event == "add":
                    sg.popup("Thêm Học Sinh Bị Lỗi Vào Danh Sách")
                if help_event == "sap_xep":
                    sg.popup("Sắp Xếp Lại Các Lớp Theo Thứ Tự Từ Thấp Đến Cao")
                if help_event == "xu_ly":
                    sg.popup("Xử Lý Các Thông Tin Nhập Từ Máy Tính Và Phân Loại Nó")
                if help_event == "luu_lai":
                    sg.popup("Lưu Lại Vào Cơ Sở Dữ Liệu Sau Khi Đã Xử Lý Hoàn Tất")
                if help_event == "xoa_hang":
                    sg.popup("Xóa Vào Hàng Muốn Chọn ( Click Vào Hàng Muốn Xóa Trên Màn )")
                if help_event == "xoa_tatca":
                    sg.popup("Xóa Tất Cả Các Thông Tin Học Sinh/Lỗi Vi Phạm Hiển Thị Trên Màn Hình")
                if help_event == "ds_loivp":
                    sg.popup("Hiển Thị DS Lỗi Vi Phạm Đã Tổng Hợp Lại Về Một Tuần Nào Đó Được Lưu Lại Trong CSDL")
                if help_event == "ds_xeploai":
                    sg.popup("Hiển Thị DS Xếp Loại Đã Tổng Hợp Lại Về Một Tuần Nào Đó Được Lưu Lại Trong CSDL")
                if help_event == "tacgia":
                    sg.popup("Tác Giả, Liên Hệ Xử Lý Khi Phầm Mềm Gặp Vẫn Đề")
                if help_event == "gui_mail":
                    sg.popup("Gửi Về Mail của các Giáo Viên Trong Nhà Trường")
            help_window.close()
                    
        if event == "sendmail":
            layout = [
                [sg.Text("Gửi Tới Mail Các GVCN",justification="center",size=(56),text_color="yellow",background_color="black")],
                [sg.Text("Nhập Tiêu Đề", size=(15, 1)), sg.InputText(key="tieude")],
                [sg.Text("Nhập Nội Dung",size=(15,5)),sg.Multiline(key="noidung",size=(43,5))],
                [sg.Text("File Đình Kèm 1",size=(15,1)),sg.Combo(docfile(),size=(43, 100),key="file1")],
                [sg.Text("File Đình Kèm 2",size=(15,1)),sg.Combo(docfile(),size=(43, 100),key="file2")],
                [sg.Text("File Đình Kèm 3",size=(15,1)),sg.Combo(docfile(),size=(43, 100),key="file3")],
                [sg.Text("File Đình Kèm 4",size=(15,1)),sg.Combo(docfile(),size=(43, 100),key="file4")],
                [sg.Button("Gửi",key="gui"),sg.Button("Huỷ",key="huy")]
            ]
            new_window = sg.Window("Gửi Tới Mail Các GVCN",layout,icon="logo.ico")
            while True:
                new_event,value_stt = new_window.read()
                if new_event in (sg.WINDOW_CLOSED, "huy"):
                    break
                else:
                    filedinhkem = [value_stt["file1"],value_stt["file2"],value_stt["file3"],value_stt["file4"]]
                    new_filedinhkem = []
                    for i in filedinhkem:
                        if i != "":
                            new_filedinhkem.append(f"{dong1}/"+i)
                    subject = value_stt["tieude"]
                    content = value_stt["noidung"]
                    sendmail(subject,content,new_filedinhkem)
                    sg.popup("Gửi Thành Công",text_color="black",icon="logo.ico")
            new_window.close()
        if event == "Save":
            layout = [
                [sg.Text("Nhập Tuần",size=(10,1)),sg.InputText(key="tuan")],
                [sg.Button("Save",key="save",size=(24),button_color="green"),
                sg.Button("Thoát",key="Thoat",size=(24),button_color="red")]
            ]
            save = sg.Window("Lưu File",layout,icon="logo.ico")
            while True:
                thamsokey, giatri = save.read()
                if thamsokey == "save":
                    df = pd.read_excel(rf"{dong2}")
                    df1 = pd.read_excel(rf"{dong3}")
                    tuan = giatri["tuan"]
                    if tuan == "":
                        sg.popup("Vui Lòng Không Để Trống Thông Tin.",icon="logo.ico")
                    else:
                        savefile(df,tuan)
                        savefilexeploai(df1,tuan)
                        khoangcach(dong2)
                        khoangcach(dong3)
                if thamsokey in (sg.WINDOW_CLOSED, "Thoat"):
                    break
            save.close()

        if event == "find":
            layout_find = [
                [sg.Text("Nhập Tuần Cần Truy Cập", size=(20, 1)), sg.InputText(key="tuan")],
                [sg.Button("Truy Cập Vào", key="truycap_ds", size=(15), button_color="green"),
                sg.Button("Thoát", key="exit_ds", size=(15), button_color="red")]
            ]
            window_find = sg.Window("Truy Cập Danh Sách Học Sinh Vi Phạm Lỗi", layout_find, icon="logo.ico")
            while True:
                event_find, values_find = window_find.read()
                if event_find == "truycap_ds":
                    tuan = "Tuần " + values_find["tuan"]
                    timkiem(dong2,tuan)
                if event_find in (sg.WINDOW_CLOSED, "exit_ds"):
                    break
            window_find.close()
                
        if event == "find_xeploai":
            layout_find = [
                [sg.Text("Nhập Tuần Cần Truy Cập", size=(20, 1)), sg.InputText(key="tuan")],
                [sg.Button("Truy Cập Vào", key="truycap_xeploai", size=(15), button_color="green"),
                sg.Button("Thoát", key="exit_xeploai", size=(15), button_color="red")]
            ]
            window_find = sg.Window("Truy Cập Danh Sách Xếp Loại Lớp", layout_find, icon="logo.ico")
            while True:
                event_find, values_find = window_find.read()
                if event_find == "truycap_xeploai":
                    tuan = "Tuần " + values_find["tuan"]
                    timkiem(dong3,tuan)
                if event_find in (sg.WINDOW_CLOSED, "exit_xeploai"):
                    break
            window_find.close()
if new_public and new_public[0] == "Activate":
    sg.popup("Phần Mềm Đang Ở Trạng Thái Public",icon="logo.ico")
    start()
else:
    try:
        seri.index(my_seri)
        sg.popup("Thiết Bị Đã Được Kích Hoạt",icon="logo.ico")
        start()
    except:
        sg.popup("Thiết Bị Chưa Được Kích Hoạt, Liên Hệ ADMIN Để Kích Hoạt Ngay",icon="logo.ico")